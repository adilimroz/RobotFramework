from openpyxl import load_workbook


def select_sheet(excel_sheet_path, sheet_name):
    workbook = load_workbook(excel_sheet_path)
    sheets = workbook.sheetnames
    i = 0
    while i < len(sheets):
        if sheets[i] == sheet_name:
            workbook.active = i
            print workbook.get_active_sheet()
            break
        else:
            i += 1
            if i == len(sheets):
                raise AssertionError("Invalid Sheet name.")


def get_single_row_data(excel_sheet_path, sheet_name, unique_key):

    workbook = load_workbook(excel_sheet_path)
    list_for_data = []
    worksheet = workbook[sheet_name]
    unique_key_col = worksheet['A']
    i = 0
    row_count = worksheet.max_row
    # print "Total row count = ", row_count

    for each_col in unique_key_col:
        if each_col.value == unique_key:
            required_value_from_col = each_col
            required_row = worksheet[required_value_from_col.row]

            for required_row_data in required_row:
                list_for_data.append(required_row_data.value)
            break
        else:
            i += 1
        print i

        if i == row_count:
            raise AssertionError("Invalid key.")

    for i in range(len(list_for_data)):
        if i == 0:
            pass
        else:
            return list_for_data[i]